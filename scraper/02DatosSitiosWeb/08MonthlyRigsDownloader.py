import os
import requests
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from urllib.parse import urljoin
import pandas as pd
from openpyxl import load_workbook
import tempfile

# Direction for download
main_url = "https://bakerhughesrigcount.gcs-web.com/intl-rig-count"

# Output Direction for the
OUTPUT_DIR = "scraper/02DatosSitiosWeb/csv"

# Function to Extract Monthly Table from the xlsx file
def extract_table_to_csv(excel_path, sheet_name="WW Monthly", table_name="WW_Monthly", output_csv=os.path.join(OUTPUT_DIR, "WorldWide_Rig_Count.csv"
)):
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb[sheet_name]  # Load the correct worksheet

        if table_name not in ws.tables:
            print(f"❌ Table '{table_name}' not found in worksheet '{sheet_name}'.")
            return None

        table = ws.tables[table_name]
        ref = table.ref  # e.g., "B7:H25"
        data = ws[ref]

        # Extract headers and rows
        headers = [cell.value for cell in data[0]]
        rows = [[cell.value for cell in row] for row in data[1:]]

        # Convert to DataFrame and save to CSV
        df = pd.DataFrame(rows, columns=headers)
        df.to_csv(output_csv, index=False)

        print(f"✅ Table '{table_name}' saved to: {output_csv}")
        return output_csv
    except Exception as e:
        print("❌ Error while processing the Excel file:", e)
        return None

#Function to Extract the Excel file
def extract_and_download_excel(base_url):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
        ))
        page = context.new_page()

        print(f"⏳ Abriendo página: {base_url}")
        page.goto(base_url, wait_until="load", timeout=60000)
        page.wait_for_selector("a", timeout=10000)

        html = page.content()
        browser.close()

    # Buscar el enlace al archivo
    soup = BeautifulSoup(html, "html.parser")
    link = soup.find("a", {
        "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "title": lambda t: t and "WorldWide Rig Count" in t
    })

    if not link or not link.get("href"):
        print("❌ Report link not found.")
        return

    # Construir la URL completa
    download_url = urljoin(base_url, link["href"])
    print("✅ Archivo encontrado:", download_url)

    # Obtener el nombre del archivo desde el título
    filename = "WorldWide_Rig_Count.xlsx"

    # Crear carpeta si no existe
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    file_path = os.path.join(OUTPUT_DIR, filename)

    # Descargar archivo
    print("⬇️ Descargando archivo...")
    response = requests.get(download_url)
    if response.status_code == 200:
        with open(file_path, "wb") as f:
            f.write(response.content)
        print(f"✅ Archivo guardado en: {file_path}")
        extract_table_to_csv(file_path)

    else:
        print(f"❌ Error al descargar el archivo: {response.status_code}")

if __name__ == "__main__":
    extract_and_download_excel(main_url)
